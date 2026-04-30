"""
Excel Utilities Module

This module provides utility functions for reading, writing, and formatting Excel files
in the CredKart automation framework. It enables data-driven testing by managing
test data stored in Excel spreadsheets and allows test result reporting through Excel formatting.

Key Functions:
    - Read cell data from Excel worksheets
    - Write test results back to Excel cells
    - Get worksheet dimensions (row and column counts)
    - Apply color formatting to cells (pass=green, fail=red)

Dependencies:
    - openpyxl: For Excel file manipulation
    - openpyxl.styles: For cell formatting and coloring

Usage Example:
    # >>> from utilities.excelutils import read_data, write_data, green_color
    # >>> # Read test data
    # >>> username = read_data('test_data.xlsx', 'LoginData', 2, 1)
    # >>> password = read_data('test_data.xlsx', 'LoginData', 2, 2)
    # >>> # Run test
    # >>> result = test_login(username, password)
    # >>> # Write result and format
    # >>> write_data('test_data.xlsx', 'LoginData', 2, 3, result)
    # >>> if result == 'Pass':
    # >>>     green_color('test_data.xlsx', 'LoginData', 2, 3)
    # >>> else:
    # >>>     red_color('test_data.xlsx', 'LoginData', 2, 3)

Excel File Organization:
    Typically stored in ./test_data/ directory
    Sheet structure:
        Row 1: Headers (Username, Password, Expected Result, etc.)
        Row 2+: Test data rows with values and result columns

Color Codes:
    - Green: #79DB77 (Success/Pass)
    - Red: #F76143 (Failure/Fail)
"""

import openpyxl
from openpyxl.styles import PatternFill


def total_rows(file_name, sheet_name):
    """
    Gets the total number of rows in an Excel worksheet.
    
    Includes header row in the count, so if worksheet has 5 data rows,
    this will return 6 (1 header + 5 data rows).
    
    Args:
        file_name (str): Path to Excel file (e.g., 'test_data/login_data.xlsx')
        sheet_name (str): Name of the worksheet to query
        
    Returns:
        int: Maximum row number (which equals total row count when starting from row 1)
        
    Example:
        # >>> rows = total_rows('test_data/login_tests.xlsx', 'LoginCredentials')
        # >>> # Returns: 11 (header + 10 data rows)
        
    Note:
        Row numbering starts at 1 (row 1 is the first row)
    """
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]

    return sheet.max_row


def total_columns(file_name, sheet_name):
    """
    Gets the total number of columns in an Excel worksheet.
    
    Args:
        file_name (str): Path to Excel file
        sheet_name (str): Name of the worksheet to query
        
    Returns:
        int: Maximum column number (which equals total column count when starting from column 1)
        
    Example:
        # >>> cols = total_columns('test_data/login_tests.xlsx', 'LoginCredentials')
        # >>> # Returns: 4 (Columns A, B, C, D)
        
    Note:
        Column numbering starts at 1 (column 1 is 'A', column 2 is 'B', etc.)
    """
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]

    return sheet.max_column


def read_data(file_name, sheet_name, row_number, column_number):
    """
    Reads a single cell value from an Excel worksheet.
    
    Args:
        file_name (str): Path to Excel file
        sheet_name (str): Name of the worksheet containing the data
        row_number (int): Row index (1-based, row 1 is the first row)
        column_number (int): Column index (1-based, column 1 is 'A')
        
    Returns:
        Any: Cell value (could be string, int, float, datetime, or None if empty)
        
    Example:
        # >>> # Read username from row 2, column 1
        # >>> username = read_data('test_data.xlsx', 'LoginData', 2, 1)
        # >>> # Returns: 'john_doe@gmail.com'
        # >>>
        # >>> # Read password from row 2, column 2
        # >>> password = read_data('test_data.xlsx', 'LoginData', 2, 2)
        # >>> # Returns: 'SecurePass123!'
        
    Note:
        Workbook is opened in read mode; changes will not be saved automatically.
        Use write_data() to persist changes to the file.
    """
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]

    return sheet.cell(row_number, column_number).value


def write_data(file_name, sheet_name, row_number, column_number, data):
    """
    Writes data to a single cell in an Excel worksheet.
    
    Args:
        file_name (str): Path to Excel file
        sheet_name (str): Name of the worksheet to write to
        row_number (int): Row index (1-based)
        column_number (int): Column index (1-based)
        data (Any): Value to write to the cell (string, int, float, datetime, etc.)
        
    Returns:
        None
        
    Example:
        # >>> # Write test result to cell at row 2, column 3
        # >>> write_data('test_data.xlsx', 'LoginData', 2, 3, 'Pass')
        # >>> # Excel cell now contains: 'Pass'
        
    Side Effect:
        File is saved automatically after writing. Previous contents of the cell
        are overwritten if cell is not empty.
        
    Note:
        This function opens, modifies, and saves the file each time it's called.
        For multiple writes, consider opening once and saving once for performance.
    """
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    sheet.cell(row_number, column_number).value = data
    workbook.save(file_name)


def green_color(file_name, sheet_name, row_number, column_number):
    """
    Applies green background color to a cell (indicating success/pass status).
    
    Color: #79DB77 (light green)
    
    Args:
        file_name (str): Path to Excel file
        sheet_name (str): Name of the worksheet containing the cell
        row_number (int): Row index (1-based)
        column_number (int): Column index (1-based)
        
    Returns:
        None
        
    Example:
        # >>> # Mark test result cell as passed (green)
        # >>> green_color('test_data.xlsx', 'LoginData', 2, 3)
        
    Side Effect:
        File is saved automatically. Cell background color is changed to green.
        Both formatting and content are preserved.
        
    Use Case:
        Typically called after a test passes:
        # >>> if test_result == 'Pass':
        # >>>     write_data(file_name, sheet, row, col, 'Pass')
        # >>>     green_color(file_name, sheet, row, col)
    """
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    # Create green fill pattern with specified color code
    fill_green = PatternFill(start_color='79DB77', end_color='79DB77', fill_type='solid')
    sheet.cell(row_number, column_number).fill = fill_green
    workbook.save(file_name)


def red_color(file_name, sheet_name, row_number, column_number):
    """
    Applies red background color to a cell (indicating failure/fail status).
    
    Color: #F76143 (orange-red)
    
    Args:
        file_name (str): Path to Excel file
        sheet_name (str): Name of the worksheet containing the cell
        row_number (int): Row index (1-based)
        column_number (int): Column index (1-based)
        
    Returns:
        None
        
    Example:
        # >>> # Mark test result cell as failed (red)
        # >>> red_color('test_data.xlsx', 'LoginData', 2, 3)
        
    Side Effect:
        File is saved automatically. Cell background color is changed to red.
        Both formatting and content are preserved.
        
    Use Case:
        Typically called after a test fails:
        # >>> if test_result == 'Fail':
        # >>>     write_data(file_name, sheet, row, col, 'Fail')
        # >>>     red_color(file_name, sheet, row, col)
    """
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    # Create red fill pattern with specified color code
    fill_red = PatternFill(start_color='F76143', end_color='F76143', fill_type='solid')
    sheet.cell(row_number, column_number).fill = fill_red
    workbook.save(file_name)
